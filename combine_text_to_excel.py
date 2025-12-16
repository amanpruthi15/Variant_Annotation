# Combined all .txt files in one excel report

import os
import sys
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Check if the correct number of command-line arguments is provided
if len(sys.argv) != 4:
    print("Usage: python combine_text_to_excel.py working_directory output_file.xlsx file_extension")
    sys.exit(1)

# Set working directory, output file name, and file extension
working_directory = sys.argv[1]
output_file = sys.argv[2]
file_extension = sys.argv[3]

# Change the working directory to the specified path
os.chdir(working_directory)

# Get a list of all text files with the specified extension in the working directory
text_files = [f for f in os.listdir() if f.endswith("." + file_extension)]

# Check if there are any text files to process
if not text_files:
    print("No {} files found in the working directory.".format(file_extension))
    sys.exit(1)

# Create a new Excel workbook
wb = Workbook()

# Loop through the text files and add them as sheets to the Excel workbook
for text_file in text_files:
    tab_name = os.path.splitext(text_file)[0]

    try:
        df = pd.read_csv(text_file, sep='\t', index_col=False, dtype=str)
    except pd.errors.EmptyDataError:
        # Create an empty sheet if the text file is empty
        ws = wb.create_sheet(title=tab_name)
        print("Empty sheet created for '{}'".format(tab_name))
        continue

    # Add a new sheet to the workbook
    ws = wb.create_sheet(title=tab_name)

    # Write the header and data to the sheet
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
        for c_idx, value in enumerate(row, start=1):
            # Try to convert value to a number if possible
            try:
                if isinstance(value, str) and value.isdigit():
                    value = int(value)
                elif isinstance(value, str):
                    value = float(value)
            except ValueError:
                pass
            ws.cell(row=r_idx, column=c_idx, value=value)

# Remove the default empty sheet created by openpyxl
if "Sheet" in wb.sheetnames and not wb["Sheet"].cell(row=1, column=1).value:
    wb.remove(wb["Sheet"])

# Save the Excel workbook with the specified output file name
wb.save(output_file)

print("Combined {} {} files into '{}'".format(len(text_files), file_extension, output_file))
